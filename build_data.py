"""Transform HK-dir grunndata Excel files into a compact JSON dataset.

Input files (in data/):
  nokkeltall-programniva-uh.xlsx           Program level, universities & høgskoler
  nokkeltall-institusjonsniva-uh.xlsx      Institution level, UH
  nokkeltall-utdanningsomradeniva-uh.xlsx  Field level, UH
  nokkeltall-programniva-fag.xlsx          Program level, fagskoler
  nokkeltall-institusjonsniva-fag.xlsx     Institution level, fagskoler
  nokkeltall-utdanningsomradeniva-fag.xlsx Field level, fagskoler

Output: data/grunndata.json
"""
import json
from pathlib import Path

import openpyxl

DATA = Path("data")
OUT = DATA / "grunndata.json"

FILES = {
    "uh": {
        "program": DATA / "nokkeltall-programniva-uh.xlsx",
        "institution": DATA / "nokkeltall-institusjonsniva-uh.xlsx",
        "field": DATA / "nokkeltall-utdanningsomradeniva-uh.xlsx",
    },
    "fag": {
        "program": DATA / "nokkeltall-programniva-fag.xlsx",
        "institution": DATA / "nokkeltall-institusjonsniva-fag.xlsx",
        "field": DATA / "nokkeltall-utdanningsomradeniva-fag.xlsx",
    },
}

METRIC_MAP = {
    "Alle søkere": "s",
    "Førstevalgssøkere": "fv",
    "Studieplasser": "p",
    "Førstevalgssøkere (kvinner)": "kv",
}


def parse_int(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def parse_pct(v):
    if v is None or v == "":
        return None
    s = str(v).replace("\xa0", "").replace("%", "").replace(",", ".").strip()
    if not s:
        return None
    try:
        return round(float(s), 1)
    except ValueError:
        return None


def load_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return header, rows[1:]


def years_from_header(header, start_col):
    return [int(y) for y in header[start_col:]]


def load_program(path):
    header, rows = load_sheet(path)
    years = years_from_header(header, 6)
    studies = {}
    for r in rows:
        if r[0] is None:
            continue
        inst, code, name, loc, field, metric = r[:6]
        values = r[6:]
        mkey = METRIC_MAP.get(metric)
        if not mkey:
            continue
        key = (inst, code, name, loc, field)
        study = studies.setdefault(
            key,
            {
                "i": inst,
                "c": code,
                "n": name,
                "l": loc,
                "f": field,
                "s": [None] * len(years),
                "fv": [None] * len(years),
                "p": [None] * len(years),
                "kv": [None] * len(years),
            },
        )
        parser = parse_pct if mkey == "kv" else parse_int
        study[mkey] = [parser(v) for v in values]
    return years, list(studies.values())


def load_agg(path, key_col):
    """Load institution/field-level aggregate sheets."""
    header, rows = load_sheet(path)
    years = years_from_header(header, 2)
    agg = {}
    for r in rows:
        if r[0] is None:
            continue
        name, metric = r[0], r[1]
        values = r[2:]
        mkey = METRIC_MAP.get(metric)
        if not mkey:
            continue
        bucket = agg.setdefault(
            name,
            {
                key_col: name,
                "s": [None] * len(years),
                "fv": [None] * len(years),
                "p": [None] * len(years),
                "kv": [None] * len(years),
            },
        )
        parser = parse_pct if mkey == "kv" else parse_int
        bucket[mkey] = [parser(v) for v in values]
    return years, list(agg.values())


def build_sector(program_path, inst_path, field_path):
    years, studies = load_program(program_path)
    _, institution_totals = load_agg(inst_path, "i")
    _, field_totals = load_agg(field_path, "f")

    institutions = sorted({s["i"] for s in studies} | {t["i"] for t in institution_totals})
    locations = sorted({s["l"] for s in studies})
    fields = sorted({s["f"] for s in studies} | {t["f"] for t in field_totals})

    inst_idx = {v: i for i, v in enumerate(institutions)}
    loc_idx = {v: i for i, v in enumerate(locations)}
    field_idx = {v: i for i, v in enumerate(fields)}

    compact_studies = []
    for s in studies:
        compact_studies.append(
            {
                "i": inst_idx[s["i"]],
                "l": loc_idx[s["l"]],
                "f": field_idx[s["f"]],
                "c": s["c"],
                "n": s["n"],
                "s": s["s"],
                "fv": s["fv"],
                "p": s["p"],
                "kv": s["kv"],
            }
        )
    compact_studies.sort(key=lambda x: (x["i"], x["f"], x["n"]))

    inst_totals = [None] * len(institutions)
    for t in institution_totals:
        inst_totals[inst_idx[t["i"]]] = {k: t[k] for k in ("s", "fv", "p", "kv")}

    field_totals_out = [None] * len(fields)
    for t in field_totals:
        field_totals_out[field_idx[t["f"]]] = {k: t[k] for k in ("s", "fv", "p", "kv")}

    return {
        "years": years,
        "institutions": institutions,
        "locations": locations,
        "fields": fields,
        "studies": compact_studies,
        "institutionTotals": inst_totals,
        "fieldTotals": field_totals_out,
    }


def main():
    out = {
        "meta": {
            "source": "HK-dir — Søkertall fra Samordna opptak",
            "url": "https://hkdir.no/sokertall-fra-samordna-opptak-til-nedlasting#Universitet%20og%20h%C3%B8gskoler",
            "sectors": {
                "uh": "Universitet og høgskoler",
                "fag": "Fagskoler",
            },
        },
        "sectors": {},
    }
    for sector, paths in FILES.items():
        out["sectors"][sector] = build_sector(paths["program"], paths["institution"], paths["field"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    total_studies = sum(len(s["studies"]) for s in out["sectors"].values())
    print(f"Wrote {OUT}  ({OUT.stat().st_size/1024:.1f} KB, {total_studies} studies)")
    for sector, s in out["sectors"].items():
        print(
            f"  {sector}: {len(s['studies'])} studies, "
            f"{len(s['institutions'])} institusjoner, "
            f"{len(s['locations'])} steder, "
            f"{len(s['fields'])} fagområder"
        )


if __name__ == "__main__":
    main()
